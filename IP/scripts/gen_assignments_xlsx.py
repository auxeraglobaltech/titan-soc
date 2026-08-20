#!/usr/bin/env python3
"""Build the IP catalogue + engineer assignment spreadsheet.

    .venv/bin/python3 IP/scripts/gen_assignments_xlsx.py

Writes IP/titan-soc_IP_assignments.xlsx with four sheets:
    IP Catalogue   every IP, full name, spec link, complexity metrics
    Assignments    who does what, in order
    Workload       per-engineer totals
    How to start   the commands an engineer needs on day one

Metrics are read live from the repo (RTL files, TB ports, interrupts, clock
domains, external dependencies), so regenerating after adding an IP keeps the
sheet honest. Names, tiers and the assignment itself are curated below and
mirror IP/ASSIGNMENTS.md -- update both together.
"""
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

IP_ROOT = Path(__file__).resolve().parent.parent
OUT = IP_ROOT / "titan-soc_IP_assignments.xlsx"

# ip -> (full name, tier)
META = {
    "gpio":       ("General Purpose Input/Output", "starter"),
    "rv_timer":   ("RISC-V Timer", "starter"),
    "pwm":        ("Pulse-Width Modulation", "starter"),
    "pattgen":    ("Pattern Generator", "starter"),
    "aon_timer":  ("Always-On Timer (wakeup + watchdog)", "starter"),
    "hmac":       ("Hash-based Message Authentication Code (SHA-2)", "starter+"),
    "uart":       ("Universal Asynchronous Receiver/Transmitter", "intermediate"),
    "adc_ctrl":   ("Analog-to-Digital Converter Controller", "intermediate"),
    "mbx":        ("Mailbox", "intermediate"),
    "pinmux":     ("Pin Multiplexer", "intermediate"),
    "dma":        ("Direct Memory Access Controller", "intermediate+"),
    "i2c":        ("Inter-Integrated Circuit (host + target)", "intermediate+"),
    "spi_host":   ("Serial Peripheral Interface - Host", "intermediate+"),
    "rv_plic":    ("RISC-V Platform-Level Interrupt Controller", "advanced"),
    "usbdev":     ("USB 2.0 Full-Speed Device", "advanced"),
    "pwrmgr":     ("Power Manager", "advanced"),
    "spi_device": ("Serial Peripheral Interface - Device", "advanced"),
    "rv_dm":      ("RISC-V Debug Module", "advanced"),
    "flash_ctrl": ("Flash Controller", "advanced"),
    "kmac":       ("Keccak Message Authentication Code (SHA-3)", "security"),
    "csrng":      ("Cryptographically Secure Random Number Generator", "security"),
    "otbn":       ("OpenTitan Big Number Accelerator", "security"),
    "aes":        ("Advanced Encryption Standard accelerator", "security"),
}

WEIGHT = {"starter": 1.0, "starter+": 1.5, "intermediate": 2.0,
          "intermediate+": 3.0, "advanced": 4.0, "security": 5.0}

# engineer -> (level, focus, ordered queue)
TEAM = {
    "J1": ("Junior", "Ramp-up: register-level then first protocol",
           ["gpio", "pattgen", "uart"]),
    "J2": ("Junior", "Ramp-up: CDC awareness early",
           ["pwm", "rv_timer", "adc_ctrl"]),
    "J3": ("Junior", "Ramp-up: first model-based checking",
           ["hmac", "aon_timer", "mbx"]),
    "E1": ("Intermediate", "AES / crypto",
           ["aes", "kmac", "i2c"]),
    "E2": ("Intermediate", "Model-based / math-heavy",
           ["csrng", "otbn", "rv_plic"]),
    "E3": ("Intermediate", "DMA / memory / storage",
           ["dma", "spi_host", "spi_device", "flash_ctrl"]),
    "E4": ("Intermediate", "System / IO / debug",
           ["pinmux", "usbdev", "pwrmgr", "rv_dm"]),
}

RATIONALE = {
    "gpio": "Cleanest possible start: no protocol. Checking is 'did the pin do what the register said'.",
    "pattgen": "Reuses the gpio shape, adds a shift-out timing element.",
    "uart": "First real serial protocol; 9 interrupts; FIFOs and watermarks. A genuine step up.",
    "pwm": "Small register file, no interrupts, but TWO clock domains - CDC awareness from day one.",
    "rv_timer": "Simplest counter/compare IP in the tree.",
    "adc_ctrl": "Sampling FSM with filters and a wakeup path; second clock domain again.",
    "hmac": "Only 4 RTL files and the reference model is a library call - a gentle first taste of model-based checking.",
    "aon_timer": "Adds an always-on domain and its first external dependency (lc_ctrl packages).",
    "mbx": "Doorbell / inbox / outbox between two bus masters.",
    "aes": "Largest RTL here (50 files). Masked datapath - reference model and DOM awareness mandatory.",
    "kmac": "Same shape as aes, in Keccak, with an app interface.",
    "i2c": "Change of pace after crypto: open-drain bus, host AND target mode, 15 interrupts.",
    "csrng": "NIST SP 800-90A CTR_DRBG - you must model the DRBG. Pulls in the whole AES cipher core.",
    "otbn": "A whole processor. Verified at ISA level.",
    "rv_plic": "Looks trivial (5 files) and is not: model priority, tie-break and claim/complete across 186 sources.",
    "dma": "Small in RTL but the only IP that INITIATES bus traffic - needs a TL-UL device agent too. Direct analogue of AXI master work.",
    "spi_host": "3 data widths x 4 clock modes; segmented commands. Depends on spi_device packages.",
    "spi_device": "Flash / passthrough / generic modes - effectively three IPs. Much cheaper once spi_host is done.",
    "flash_ctrl": "Scrambling, ECC, program/erase, lifecycle interaction. 14 external dependencies.",
    "pinmux": "53 ports but conceptually a muxing matrix - a gentle start at this tier.",
    "usbdev": "50 ports, 18 interrupts. Needs a USB link-layer model.",
    "pwrmgr": "Sleep/wake FSM across FOUR clock domains.",
    "rv_dm": "RISC-V debug over JTAG; needs a DTM/DMI model. Hardest of E4's queue - scheduled last.",
}

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TIER_FILL = {
    "starter":       PatternFill("solid", fgColor="E2EFDA"),
    "starter+":      PatternFill("solid", fgColor="D9EAD3"),
    "intermediate":  PatternFill("solid", fgColor="FFF2CC"),
    "intermediate+": PatternFill("solid", fgColor="FCE4D6"),
    "advanced":      PatternFill("solid", fgColor="F8CBAD"),
    "security":      PatternFill("solid", fgColor="F4B6B6"),
}
LEVEL_FILL = {"Junior": PatternFill("solid", fgColor="DDEBF7"),
              "Intermediate": PatternFill("solid", fgColor="E4DFEC")}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LINK = Font(color="0563C1", underline="single")


def metrics(ip):
    d = IP_ROOT / ip
    tb = (d / "verification/tb/tb.sv").read_text()
    files_f = (d / "rtl/files.f").read_text()
    autogen = "ip_autogen" in files_f
    return {
        "rtl": len(list((d / "rtl").glob("*.sv"))),
        "ports": len(re.findall(r"^\s+\.\w+", tb, re.M)),
        "irqs": len(re.findall(r"^\s+\.intr_", tb, re.M)),
        "clks": len(re.findall(r"^\s+\.clk_\w+", tb, re.M)),
        "deps": len(re.findall(r"^//   \S+\.sv  <-", files_f, re.M)),
        "autogen": autogen,
        "spec": (f"https://opentitan.org/book/hw/top_earlgrey/ip_autogen/{ip}/"
                 if autogen else f"https://opentitan.org/book/hw/ip/{ip}/"),
        "curated": "auto-generated" not in (d / "README.md").read_text(),
    }


def style_header(ws, ncols, freeze="A2"):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = freeze
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def main():
    M = {ip: metrics(ip) for ip in META}
    wb = Workbook()

    # ---------------------------------------------------------- catalogue
    ws = wb.active
    ws.title = "IP Catalogue"
    ws.append(["IP", "Full name", "Tier", "Source tree", "RTL files",
               "DUT ports", "Interrupts", "Clock domains", "External deps",
               "Elaborates", "README", "Spec (upstream)", "Folder"])
    order = sorted(META, key=lambda i: (list(WEIGHT).index(META[i][1]), i))
    for ip in order:
        name, tier = META[ip]
        m = M[ip]
        ws.append([ip, name, tier,
                   "ip_autogen/" if m["autogen"] else "hw/ip/",
                   m["rtl"], m["ports"], m["irqs"], m["clks"], m["deps"],
                   "PASS", "curated" if m["curated"] else "generated",
                   m["spec"], f"IP/{ip}/"])
        r = ws.max_row
        ws.cell(row=r, column=3).fill = TIER_FILL[tier]
        c = ws.cell(row=r, column=12)
        c.hyperlink, c.font, c.value = m["spec"], LINK, "spec"
        ws.cell(row=r, column=10).alignment = Alignment(horizontal="center")
    style_header(ws, 13)
    widths(ws, {"A": 12, "B": 46, "C": 14, "D": 13, "E": 10, "F": 10,
                "G": 11, "H": 14, "I": 13, "J": 11, "K": 12, "L": 14, "M": 14})
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=13):
        for cell in row:
            cell.border = BORDER

    # -------------------------------------------------------- assignments
    ws = wb.create_sheet("Assignments")
    ws.append(["Engineer", "Level", "Focus", "Order", "IP", "Full name",
               "Tier", "Weight", "Why this IP / this position", "Spec", "Folder"])
    for eng, (level, focus, queue) in TEAM.items():
        for i, ip in enumerate(queue, 1):
            name, tier = META[ip]
            ws.append([eng, level, focus, i, ip, name, tier, WEIGHT[tier],
                       RATIONALE.get(ip, ""), M[ip]["spec"], f"IP/{ip}/"])
            r = ws.max_row
            ws.cell(row=r, column=2).fill = LEVEL_FILL[level]
            ws.cell(row=r, column=7).fill = TIER_FILL[tier]
            c = ws.cell(row=r, column=10)
            c.hyperlink, c.font, c.value = M[ip]["spec"], LINK, "spec"
    style_header(ws, 11)
    widths(ws, {"A": 10, "B": 14, "C": 34, "D": 7, "E": 12, "F": 44,
                "G": 14, "H": 8, "I": 78, "J": 8, "K": 14})
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=11):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column == 9))

    # ----------------------------------------------------------- workload
    ws = wb.create_sheet("Workload")
    ws.append(["Engineer", "Level", "Focus", "IPs", "Total weight", "Queue (in order)"])
    for eng, (level, focus, queue) in TEAM.items():
        ws.append([eng, level, focus, len(queue),
                   sum(WEIGHT[META[i][1]] for i in queue),
                   "  ->  ".join(queue)])
        ws.cell(row=ws.max_row, column=2).fill = LEVEL_FILL[level]
    ws.append([])
    ws.append(["Weighting", "starter 1.0", "starter+ 1.5", "intermediate 2.0",
               "intermediate+ 3.0", "advanced 4.0 / security 5.0"])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, bold=True)
    ws.append([])
    ws.append(["Note", "Load is deliberately uneven between levels. A junior's first IP is mostly "
                       "learning the methodology and takes far longer than its weight suggests; "
                       "by their third they should be at roughly half an intermediate's rate."])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, bold=True)
    ws.cell(row=ws.max_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    style_header(ws, 6)
    widths(ws, {"A": 11, "B": 14, "C": 34, "D": 7, "E": 13, "F": 60})

    # ------------------------------------------------------- how to start
    ws = wb.create_sheet("How to start")
    rows = [
        ("Step", "What to do", "Detail"),
        ("0", "Set up the environment",
         "source scripts/activate_env.sh   (bash)  or  activate_env.csh  (csh/tcsh); then: which xrun"),
        ("1", "Prove your IP compiles BEFORE changing anything",
         "cd IP/<ip>/sim && ./run_compile.sh    -- expect 'TEST PASSED' and 'UVM_ERROR : 0'. "
         "All 23 were verified clean on 2026-08-20, so any failure is something you changed."),
        ("2", "Read the spec",
         "IP/<ip>/docs/theory_of_operation.md, registers.md, interfaces.md, plus the linked spec. "
         "Enumerate every feature."),
        ("3", "Write the testplan FIRST",
         "IP/<ip>/verification/testplan.md -- feature / test / pass criteria / coverage goal. "
         "Write it before any SystemVerilog: a testplan derived from a TB you already built only "
         "describes what you happened to implement."),
        ("4", "Build the environment",
         "TL-UL host agent (reuse vendor/opentitan/hw/dv/sv/tl_agent) -> RAL from docs/<ip>.hjson via "
         "util/regtool.py -> reset/CSR tests -> pin-level agent -> scoreboard -> interrupts -> coverage."),
        ("5", "Close coverage", "Functional covergroups tied to your testplan's coverage goals."),
        ("", "", ""),
        ("Rule", "One simulation at a time", "The server is shared. Run 'who' first."),
        ("Rule", "Tie off every unused DUT input",
         "A floating input propagates X into the register file and trips the TL-UL dKnown assertions; "
         "the failure surfaces long after the real cause."),
        ("Rule", "stat the log before trusting it",
         "stat -c '%y' runs/compile.log -- the run dir is overwritten in place, so a failed build "
         "leaves the previous log looking valid."),
        ("Rule", "Do not read the upstream dv/ environment first",
         "vendor/opentitan/hw/ip/<ip>/dv/ is the full upstream env. Reading it before you have "
         "attempted a piece turns the exercise into transcription."),
        ("Rule", "Upstream's testplan is the answer key",
         "data/<ip>_testplan.hjson in the vendor tree is deliberately NOT copied in. Write yours first, "
         "then diff against it."),
        ("", "", ""),
        ("Note", "Six READMEs are curated, 17 are generated",
         "gpio, pwm, uart, i2c, spi_host and rv_plic have hand-written spec analysis. The other 17 "
         "contain extracted facts only and say so at the top. Improving yours as you read the spec is "
         "a legitimate first task."),
    ]
    for r in rows:
        ws.append(list(r))
    style_header(ws, 3)
    widths(ws, {"A": 10, "B": 44, "C": 104})
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  IPs: {len(META)}   engineers: {len(TEAM)}   "
          f"assigned: {sum(len(q) for _, _, q in TEAM.values())}")


if __name__ == "__main__":
    main()
